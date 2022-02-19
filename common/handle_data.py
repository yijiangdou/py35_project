import openpyxl

class HandleExcel:
    def __init__(self,filename,sheetname):
        self.filename = filename
        self.sheetname = sheetname
    def read_data(self):
        """读取excel数据"""
        wk = openpyxl.load_workbook(self.filename)
        sh = wk[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行之外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title,data))
            cases.append(dic)
        return cases
    def write_data(self,row,column,value):
        """数据写入"""
        wk = openpyxl.load_workbook(self.filename)
        sh = wk[self.sheetname]
        sh.cell(row=row,column=column,value=value)
        wk.save(self.filename)